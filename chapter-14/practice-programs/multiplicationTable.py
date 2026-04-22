#! python3
# multiplicationTable.py - Creates an Excel Spreadsheet with an N * N multiplication table

import sys, openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

tableMaxVal = int(sys.argv[1])

wb = openpyxl.Workbook()

sheet = wb.active

for i in range(1, tableMaxVal + 1):
    sheet['A' + str(1 + i)] = i
    sheet[get_column_letter(1 + i) + str(1)] = i
    sheet['A' + str(1 + i)].font = Font(bold=True)
    sheet[get_column_letter(1 + i) + str(1)].font = Font(bold=True)

for rowofCellObjects in sheet['B2':get_column_letter(tableMaxVal + 1)+str(tableMaxVal + 1)]:
    for rowCell in enumerate(rowofCellObjects):
        sheet[rowCell[1].coordinate] = f"={get_column_letter(sheet[rowCell[1].coordinate].column) + str(1)}*{'A' + str(sheet[rowCell[1].coordinate].row)}"

wb.save('multiplicationTable.xlsx')
