#! python3
# blankRowInserter.py - Inserts M blank rows starting at N row

import sys, openpyxl
from openpyxl.utils import get_column_letter

blankRowNum = int(sys.argv[1])
numRowsBlank = int(sys.argv[2])
insertSpreadsheet = sys.argv[3]

wb = openpyxl.load_workbook(insertSpreadsheet)
beforeSheet = wb[wb.sheetnames[0]]
beforeSheet.title = 'before'

wb.create_sheet(title="after")
afterSheet = wb['after']

for rowNum, row in enumerate(beforeSheet['A1': get_column_letter(beforeSheet.max_column) + str(beforeSheet.max_row)]):
    if rowNum+1 >= blankRowNum:
        for rowCell in row:
            afterSheet[get_column_letter(beforeSheet[rowCell.coordinate].column) + str(int(beforeSheet[rowCell.coordinate].row) + numRowsBlank)] = beforeSheet[rowCell.coordinate].value
    else:
        for rowCell in row:
            afterSheet[rowCell.coordinate] = beforeSheet[rowCell.coordinate].value

wb.save('insertedVals.xlsx')
